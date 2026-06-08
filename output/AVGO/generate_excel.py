#!/usr/bin/env python3
"""
AVGO Valuation Excel Model — all valuation math in a structured workbook.
Sheets: Cover | Market Stats | DCF | Relative | SOTP | Sensitivity | Scenarios | Peers
"""

import openpyxl as xl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

OUT = "/home/user/ERFS/output/AVGO/AVGO_valuation.xlsx"

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_):  return PatternFill("solid", fgColor=hex_.lstrip("#"))
def font(bold=False, italic=False, size=10, color="000000", name="Calibri"):
    return Font(bold=bold, italic=italic, size=size, color=color.lstrip("#"), name=name)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="D0D5E0")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    s = Side(style="medium", color="0D1B3E")
    return Border(bottom=s)

# Palettes
NAV = "0D1B3E"; BLU = "1A3A8F"; GRN = "1B7A4A"; RED = "C0392B"
AMB = "E67E22"; LGR = "F4F6FA"; MGR = "D0D5E0"; WHT = "FFFFFF"
YEL = "FFF9C4"; GLD = "FFD700"; CYN = "E3F2FD"

def hdr(ws, row, col, val, bg=BLU, fg=WHT, bold=True, sz=10, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, size=sz)
    c.alignment = align(halign, "center")
    c.border = thin_border()
    return c

def cell(ws, row, col, val, bg=WHT, bold=False, fg="000000", halign="right",
         num_fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, italic=italic)
    c.alignment = align(halign, "center")
    c.border = thin_border()
    if num_fmt: c.number_format = num_fmt
    return c

def label(ws, row, col, val, bg=LGR, bold=False):
    return cell(ws, row, col, val, bg=bg, bold=bold, halign="left")

def money(ws, row, col, val, bg=WHT, bold=False, fg="000000"):
    return cell(ws, row, col, val, bg=bg, bold=bold, fg=fg, num_fmt='#,##0.00')

def pct(ws, row, col, val, bg=WHT, bold=False):
    return cell(ws, row, col, val, bg=bg, bold=bold, num_fmt='0.0%')

def merge_hdr(ws, row, c1, c2, val, bg=NAV, fg=WHT, sz=12):
    c = ws.cell(row=row, column=c1, value=val)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c.fill = fill(bg); c.font = font(bold=True, color=fg, size=sz)
    c.alignment = align("center","center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = xl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 1: Cover
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cover")
set_col_widths(ws, [4, 30, 25, 25, 25, 4])

merge_hdr(ws, 1, 1, 6, "BROADCOM INC. (AVGO) — VALUATION MODEL", bg=NAV, sz=16)
merge_hdr(ws, 2, 1, 6, "Institutional Equity Research  |  Date: 2026-06-08  |  Data: yfinance 1.4.1", bg=NAV, sz=10)

ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 18

r = 4
meta = [
    ("Rating", "OVERWEIGHT", "Price Target", "$475"),
    ("Current Price", "$385.73", "Upside to Target", "+23.2%"),
    ("Expected Value (PW)", "$445", "Analyst Consensus Mean", "$517.61"),
    ("Market Cap", "$1,826B", "EV", "$1,876B"),
    ("TTM Revenue", "$75.5B", "TTM EBITDA", "$41.9B"),
    ("FY2025 FCF", "$26.9B (42.1% margin)", "Net Debt (Q1 FY26)", "$51.9B"),
    ("Fwd P/E (nGAAP)", "20.0×", "EV/EBITDA (TTM)", "44.7×"),
    ("ADTV (30d)", "$11.6B", "30d Realized Vol", "69.4% ann."),
    ("Beta", "1.433", "Short % Float", "1.15%"),
    ("WACC (Adjudicated)", "12.18%", "ke (CAPM)", "12.42%"),
    ("DCF Base Implied", "$218 (Gordon TV)", "DCF Base+ Implied", "$304 (blended TV)"),
    ("Relative Implied", "$252–$464 (range)", "SOTP Implied", "$189"),
]
for k1, v1, k2, v2 in meta:
    label(ws, r, 2, k1, bg=LGR, bold=True)
    cell(ws, r, 3, v1, halign="left")
    label(ws, r, 4, k2, bg=LGR, bold=True)
    cell(ws, r, 5, v2, halign="left")
    r += 1

r += 1
merge_hdr(ws, r, 2, 5, "Research Cycle: Alpha · Charlie (bull) · Kilo (bear) · Delta (audit)", bg=BLU, sz=9)
r += 1
merge_hdr(ws, r, 2, 5, "Not financial advice. For professional investors only.", bg="555E70", sz=9)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 2: Market Stats
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Market Stats")
set_col_widths(ws, [4, 35, 20, 35, 20, 4])
freeze(ws, "B2")
merge_hdr(ws, 1, 1, 6, "AVGO — Market Statistics (yfinance, 2026-06-08)", bg=NAV, sz=12)

rows = [
    ("Current Price", "$385.73", "Analyst Mean Target", "$517.61 (+34%)"),
    ("Market Cap", "$1,826B", "Analyst Median Target", "$525.00 (+36%)"),
    ("Enterprise Value", "$1,876B", "Analyst High / Low", "$650 / $215.88"),
    ("Shares Outstanding", "4.735B", "# Analysts", "45"),
    ("Float Shares", "4.688B", "Mean Rating", "1.33 / 5 (Strong Buy)"),
    ("52-Week High", "$495.00", "Beta (5yr monthly)", "1.433"),
    ("52-Week Low", "$241.11", "Sector", "Technology"),
    ("Drawdown from 52W High", "−22.1%", "Industry", "Semiconductors"),
    ("YTD Return (2026)", "+11.2%", "Fiscal Year End", "October 31"),
    ("1-Year Return", "+57.5%", "Employees", "~40,000"),
    ("2-Year Return", "+179.8%", "HQ", "Palo Alto, CA"),
    ("ADTV (30-day, USD)", "$11.6B", "Listing", "NASDAQ"),
    ("30d Realized Volatility", "69.4% ann.", "Trailing EPS (GAAP)", "$6.02"),
    ("Short % of Float", "1.15%", "Forward EPS (nGAAP)", "$19.32 (FY2027E)"),
    ("Short Ratio (days)", "2.7 days", "Trailing P/E (GAAP)", "64.1×"),
    ("Insider Ownership", "1.95%", "Forward P/E (nGAAP)", "20.0×"),
    ("Institutional Ownership", "80.3%", "EV/Revenue (TTM)", "24.9×"),
    ("Dividend / Yield", "$2.60 / 0.67%", "EV/EBITDA (TTM)", "44.7×"),
]
r = 2
for k1, v1, k2, v2 in rows:
    label(ws, r, 2, k1, bold=True)
    cell(ws, r, 3, v1, halign="left")
    label(ws, r, 4, k2, bold=True)
    cell(ws, r, 5, v2, halign="left")
    r += 1

r += 1
merge_hdr(ws, r, 2, 5, "Source: yfinance 1.4.1 — info, price history, analyst_price_targets, recommendations", bg=BLU, sz=9)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 3: Historicals
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Historicals")
set_col_widths(ws, [4, 32, 16, 16, 16, 16, 16, 16, 16, 4])
freeze(ws, "C3")
merge_hdr(ws, 1, 1, 9, "AVGO — Historical Financials (USD Billions unless noted)", bg=NAV, sz=12)

periods = ["FY2022", "FY2023", "FY2024", "FY2025", "TTM", "FY2026E", "FY2027E"]
for i, p in enumerate(periods):
    bg_ = CYN if "E" in p else (LGR if p == "TTM" else WHT)
    hdr(ws, 2, 3+i, p, bg=BLU if "E" not in p and p != "TTM" else AMB)

data = [
    # Label,  FY22,  FY23,  FY24,  FY25,  TTM,     FY26E,   FY27E
    ("Revenue ($B)",           33.2,  35.8,  51.6,  63.9,  75.5,   106.0,  170.5),
    ("YoY Revenue Growth",     0.207, 0.079, 0.440, 0.239, 0.479,  0.659,  0.608),
    ("Gross Profit ($B)",      22.1,  24.7,  32.5,  43.3,  57.6,   None,   None),
    ("Gross Margin",           0.666, 0.689, 0.630, 0.678, 0.763,  None,   None),
    ("Operating Income ($B)",  14.3,  16.5,  15.0,  26.1,  37.0,   None,   None),
    ("Operating Margin",       0.430, 0.459, 0.291, 0.408, 0.490,  None,   None),
    ("EBITDA ($B)",            19.2,  20.6,  23.9,  34.7,  41.9,   None,   None),
    ("EBITDA Margin",          0.578, 0.575, 0.463, 0.543, 0.555,  None,   None),
    ("GAAP Net Income ($B)",   11.5,  14.1,   5.9,  23.1,  None,   None,   None),
    ("FCF ($B)",               16.3,  17.6,  19.4,  26.9,  32.0,   None,   None),
    ("FCF Margin",             0.491, 0.492, 0.376, 0.421, None,   None,   None),
    ("GAAP EPS ($)",           2.65,  3.30,  1.23,  4.77,  6.02,   None,   None),
    ("nGAAP EPS ($, est.)",    None,  None,  None,  None,  None,   11.615, 19.321),
    ("Capex ($M)",             424,   452,   548,   623,   None,   None,   None),
    ("Capex % Revenue",        0.013, 0.013, 0.011, 0.010, None,   None,   None),
    ("Total Debt ($B)",        39.5,  39.2,  67.6,  65.1,  66.1,   None,   None),
    ("Cash ($B)",              12.4,  14.2,   9.3,  16.2,  14.2,   None,   None),
    ("Net Debt ($B)",          27.1,  25.0,  58.2,  48.9,  51.9,   None,   None),
    ("Net Debt / EBITDA",       1.4,   1.2,   2.4,   1.4,   1.2,   None,   None),
    ("Interest Expense ($B)",  None,  1.622, 3.953, 3.210, None,   None,   None),
    ("Interest Coverage (×)",  None,  None,   3.6,   8.1,  None,   None,   None),
    ("FY25 Debt Repayment ($B)",None, None,  19.6,  18.5,  None,   None,   None),
]

pct_rows = {1,3,5,7,10,14}   # row offsets that are percentages
r = 3
for i, (lbl, *vals) in enumerate(data):
    label(ws, r, 2, lbl, bold=(i == 0))
    for j, v in enumerate(vals):
        if v is None:
            cell(ws, r, 3+j, "—", halign="center", bg=LGR if j == 4 else WHT)
        elif i in pct_rows:
            pct(ws, r, 3+j, v, bg=CYN if j in (5,6) else (LGR if j == 4 else WHT))
        else:
            money(ws, r, 3+j, v, bg=CYN if j in (5,6) else (LGR if j == 4 else WHT),
                  bold=(i == 0))
    r += 1

r += 1
merge_hdr(ws, r, 2, 9, "Source: yfinance income_stmt, cashflow, balance_sheet (annual). FY26E/FY27E: yfinance consensus.", bg=BLU, sz=9)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 4: WACC & DCF
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("WACC & DCF")
set_col_widths(ws, [4, 38, 20, 20, 20, 20, 20, 20, 20, 4])
freeze(ws, "B2")
merge_hdr(ws, 1, 1, 9, "AVGO — WACC Adjudication & DCF Valuation", bg=NAV, sz=12)

# WACC block
r = 3
merge_hdr(ws, r, 2, 5, "WACC INPUTS & ADJUDICATION", bg=BLU); r += 1
wacc_rows = [
    ("Risk-Free Rate (rf)", "10Y US Treasury", 0.0454),
    ("Beta (β)", "yfinance info.beta", 1.433),
    ("Equity Risk Premium (ERP)", "Damodaran US implied", 0.055),
    ("ke = rf + β × ERP", "CAPM", 0.1242),
    ("Equity weight (E/V)", "market_cap / EV", 0.972),
    ("Debt weight (D/V)", "1 − E/V", 0.028),
    ("kd (effective, pre-tax)", "interest_exp / total_debt", 0.0486),
    ("Tax rate (statutory)", "21%", 0.21),
    ("kd (after-tax)", "kd × (1 − tax)", 0.0384),
    ("Formula WACC", "E/V×ke + D/V×kd×(1−t)", 0.1218),
]
for lbl, src, val in wacc_rows:
    label(ws, r, 2, lbl, bold=("WACC" in lbl or "ke" in lbl))
    cell(ws, r, 3, src, halign="left")
    pct(ws, r, 4, val, bold=("WACC" in lbl or "ke" in lbl))
    r += 1

r += 1
merge_hdr(ws, r, 2, 5, "ADJUDICATION CROSSCHECK", bg=AMB, fg=WHT); r += 1
xcheck = [
    ("Semiconductor sector WACC band", "10.0% – 12.0%"),
    ("Formula WACC vs band", "12.18% — just ABOVE band"),
    ("Peer-implied WACC (NVDA/MRVL/QCOM)", "~10.5%"),
    ("Market-implied ke (1/fwdPE + 2.5%)", "~7.5%"),
    ("ADJUDICATED WACC", "12.18% — formula value (near-100% equity; no debt tax shield)"),
]
for lbl, note in xcheck:
    label(ws, r, 2, lbl, bold=("ADJUDICATED" in lbl))
    cell(ws, r, 3, note, halign="left", bold=("ADJUDICATED" in lbl),
         bg=YEL if "ADJUDICATED" in lbl else WHT)
    r += 1

# DCF block
r += 1
merge_hdr(ws, r, 2, 9, "DCF — 5-YEAR FCFF PROJECTION (Base: WACC=12.18%, Tax=21%, EBIT Margin=45%)", bg=BLU)
r += 1
years = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
for i, y in enumerate(years):
    hdr(ws, r, 3+i, y, bg=AMB)
cell(ws, r, 2, "Item", bg=BLU, bold=True, fg=WHT); r += 1

growth = [0.659, 0.608, 0.25, 0.15, 0.08]
rev_base = 63.887
revs = []
r_data = rev_base
for g in growth:
    r_data *= (1 + g)
    revs.append(round(r_data, 2))

ebit_m = 0.45; da_m = 0.137; capex_m = 0.011; nwc_m = 0.005; tax = 0.21

ebits   = [round(v * ebit_m, 2) for v in revs]
nopats  = [round(e * (1 - tax), 2) for e in ebits]
das     = [round(v * da_m, 2) for v in revs]
capexs  = [round(v * capex_m, 2) for v in revs]
nwcs    = [round(v * nwc_m, 2) for v in revs]
fcffs   = [round(nopats[i] + das[i] - capexs[i] - nwcs[i], 2) for i in range(5)]

wacc_ = 0.1218
pv_fcffs = [round(fcffs[i] / (1 + wacc_)**(i+1), 2) for i in range(5)]
g_term = 0.025
tv_gordon = round(fcffs[-1] * (1 + g_term) / (wacc_ - g_term), 2)
tv_exit   = round(revs[-1] * (ebit_m + da_m) * 18, 2)
tv_blend  = round((tv_gordon + tv_exit) / 2, 2)
pv_tv_gordon = round(tv_gordon / (1 + wacc_)**5, 2)
pv_tv_blend  = round(tv_blend  / (1 + wacc_)**5, 2)
pv_fcff_total= round(sum(pv_fcffs), 2)

shares = 4.7347
cash_  = 14.174; debt_  = 66.057; net_debt_ = debt_ - cash_

ev_gordon = pv_fcff_total + pv_tv_gordon
eq_gordon = ev_gordon + cash_ - debt_
px_gordon = round(eq_gordon / shares, 2)

ev_blend  = pv_fcff_total + pv_tv_blend
eq_blend  = ev_blend + cash_ - debt_
px_blend  = round(eq_blend / shares, 2)

dcf_rows = [
    ("Revenue ($B)",          revs,   False, False),
    ("YoY Growth",            growth, True,  False),
    ("EBIT ($B) [45%]",       ebits,  False, False),
    ("NOPAT ($B) [×(1−21%)]", nopats, False, False),
    ("D&A ($B) [13.7%]",      das,    False, False),
    ("Capex ($B) [1.1%]",     capexs, False, False),
    ("ΔNWC ($B) [0.5%]",      nwcs,   False, False),
    ("FCFF ($B)",              fcffs,  False, True),
    ("PV of FCFF ($B)",        pv_fcffs, False, False),
]
for lbl, vals, is_pct, bold in dcf_rows:
    label(ws, r, 2, lbl, bold=bold)
    for j, v in enumerate(vals):
        if is_pct:
            pct(ws, r, 3+j, v, bold=bold)
        else:
            money(ws, r, 3+j, v, bold=bold)
    r += 1

r += 1
merge_hdr(ws, r, 2, 9, "TERMINAL VALUE & EV BRIDGE", bg=BLU); r += 1
bridge = [
    ("Terminal growth rate (g)", "", "", "", g_term),
    ("Terminal FCFF (Year 5)", "", "", "", fcffs[-1]),
    ("TV Gordon Growth ($B)", "", "", "", tv_gordon),
    ("TV Exit Multiple (18× EBITDA, $B)", "", "", "", tv_exit),
    ("TV Blended ($B)", "", "", "", tv_blend),
    ("PV of TV — Gordon ($B)", "", "", "", pv_tv_gordon),
    ("PV of TV — Blended ($B)", "", "", "", pv_tv_blend),
    ("Sum PV of FCFFs ($B)", "", "", "", pv_fcff_total),
    ("EV — Gordon TV ($B)", "", "", "", ev_gordon),
    ("EV — Blended TV ($B)", "", "", "", ev_blend),
    ("+ Cash ($B)", "", "", "", cash_),
    ("− Total Debt ($B)", "", "", "", debt_),
    ("Equity Value — Gordon ($B)", "", "", "", eq_gordon),
    ("Equity Value — Blended ($B)", "", "", "", eq_blend),
    ("Diluted Shares (B)", "", "", "", shares),
    ("→ Implied Price — Gordon TV", "", "", "", px_gordon),
    ("→ Implied Price — Blended TV", "", "", "", px_blend),
]
for i, (lbl, _, __, ___, val) in enumerate(bridge):
    is_implied = lbl.startswith("→")
    label(ws, r, 2, lbl, bold=is_implied)
    c = money(ws, r, 6, val, bold=is_implied, bg=YEL if is_implied else WHT,
              fg=GRN if is_implied else "000000")
    if lbl.startswith("Terminal growth"):
        c.number_format = "0.0%"
    if lbl.startswith("Diluted Shares"):
        c.number_format = "#,##0.00"
    r += 1

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 5: Sensitivity
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DCF Sensitivity")
set_col_widths(ws, [4, 20, 18, 18, 18, 18, 18, 4])
merge_hdr(ws, 1, 1, 8, "DCF SENSITIVITY: WACC × Terminal g  (Gordon TV, $B equity, implied $/share)", bg=NAV, sz=12)

r = 3
merge_hdr(ws, r, 2, 8, "Implied Share Price ($) — Gordon TV only", bg=BLU); r += 1
wacc_vals = [0.105, 0.110, 0.1218, 0.130, 0.140]
g_vals    = [0.015, 0.020, 0.025, 0.030, 0.035]
hdr(ws, r, 2, "WACC \\ g")
for j, g in enumerate(g_vals):
    hdr(ws, r, 3+j, f"g = {g:.1%}")
r += 1

for wi, w in enumerate(wacc_vals):
    is_base_w = abs(w - 0.1218) < 0.001
    label(ws, r, 2, f"WACC = {w:.1%}", bold=is_base_w, bg=YEL if is_base_w else LGR)
    for gi, g in enumerate(g_vals):
        is_base_g = abs(g - 0.025) < 0.001
        is_base = is_base_w and is_base_g
        if w <= g:
            cell(ws, r, 3+gi, "n/a", halign="center", bg="FFCCCC")
        else:
            tv_g = fcffs[-1] * (1 + g) / (w - g)
            pv_tv_s = tv_g / (1 + w)**5
            ev_s = pv_fcff_total + pv_tv_s
            eq_s = ev_s + cash_ - debt_
            px_s = round(eq_s / shares, 0)

            bg_ = YEL if is_base else (CYN if is_base_g else (LGR if is_base_w else WHT))
            c = money(ws, r, 3+gi, px_s, bg=bg_, bold=is_base)
            c.number_format = '"$"#,##0'
            if px_s > 385:
                c.font = Font(bold=is_base, color=GRN, name="Calibri")
            elif px_s < 250:
                c.font = Font(bold=is_base, color=RED, name="Calibri")
    r += 1

r += 1
merge_hdr(ws, r, 2, 8, "★ = Base case (WACC 12.18%, g 2.5%) → $218  |  Current Price: $385.73  |  Green = above current, Red = <$250", bg=AMB)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 6: Relative & SOTP
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Relative & SOTP")
set_col_widths(ws, [4, 38, 22, 22, 22, 22, 4])
freeze(ws, "B2")
merge_hdr(ws, 1, 1, 7, "AVGO — Relative Valuation & SOTP", bg=NAV, sz=12)

r = 3
merge_hdr(ws, r, 2, 7, "PEER MULTIPLES (yfinance, 2026-06-08)", bg=BLU); r += 1
peer_hdrs = ["Company", "Fwd P/E", "EV/EBITDA", "EV/Revenue", "Rev Growth", "Gross Margin"]
for i, h in enumerate(peer_hdrs):
    hdr(ws, r, 2+i, h)
r += 1

peers = [
    ("AVGO ★",  20.0, 44.7, 24.9, 0.659, 0.763),
    ("NVDA",    16.2, 29.7, 19.4, 0.850, 0.740),
    ("MRVL",    42.7, 85.5, 26.6, 0.280, 0.520),
    ("QCOM",    20.2, 17.9,  5.2,-0.040, 0.550),
    ("AMD",     35.7,101.2, 20.1, 0.380, 0.530),
    ("ORCL",    26.5, 27.1, 11.6, 0.220, 0.670),
    ("MSFT",    21.5, 17.0,  9.9, 0.180, 0.680),
    ("Median",  24.0, 28.4, 15.5, None,  None),
]
for name, fpe, eveb, evr, gr, gm in peers:
    is_avgo = "AVGO" in name; is_med = name == "Median"
    bg_ = CYN if is_avgo else (LGR if is_med else WHT)
    bold_ = is_avgo or is_med
    label(ws, r, 2, name, bold=bold_, bg=bg_)
    money(ws, r, 3, fpe,  bg=bg_, bold=bold_); ws.cell(r,3).number_format = '0.0"×"'
    money(ws, r, 4, eveb, bg=bg_, bold=bold_); ws.cell(r,4).number_format = '0.0"×"'
    money(ws, r, 5, evr,  bg=bg_, bold=bold_); ws.cell(r,5).number_format = '0.0"×"'
    if gr is not None: pct(ws, r, 6, gr, bg=bg_, bold=bold_)
    else: cell(ws, r, 6, "—", halign="center", bg=bg_)
    if gm is not None: pct(ws, r, 7, gm, bg=bg_, bold=bold_)
    else: cell(ws, r, 7, "—", halign="center", bg=bg_)
    r += 1

r += 1
merge_hdr(ws, r, 2, 7, "RELATIVE VALUATION — IMPLIED PRICES", bg=BLU); r += 1

rel_hdrs = ["Method", "Metric Used", "Multiple", "Implied Price"]
for i, h in enumerate(rel_hdrs):
    hdr(ws, r, 2+i, h)
r += 1
eps_26 = 11.615; eps_27 = 19.321; ebitda_ttm = 41.9; rev_ttm = 75.5

med_pe = 24.0; med_eveb = 28.4; med_evr = 15.5

rel_methods = [
    ("Fwd P/E (FY2026E, current yr)",   f"EPS ${eps_26}",          f"{med_pe}× median",    round(med_pe * eps_26, 2)),
    ("Fwd P/E (FY2027E, NTM)",          f"EPS ${eps_27}",          f"{med_pe}× median",    round(med_pe * eps_27, 2)),
    ("EV/EBITDA (TTM)",                 f"EBITDA ${ebitda_ttm}B",  f"{med_eveb}× median",  round((med_eveb * ebitda_ttm - net_debt_) / shares, 2)),
    ("EV/Revenue (TTM)",                f"Revenue ${rev_ttm}B",    f"{med_evr}× median",   round((med_evr * rev_ttm - net_debt_) / shares, 2)),
    ("→ Blended Relative (current yr)", "P/E + EV/EBITDA + EV/Rev","average",              round(((med_pe * eps_26) + (med_eveb * ebitda_ttm - net_debt_)/shares + (med_evr * rev_ttm - net_debt_)/shares) / 3, 2)),
    ("→ Adopted Target",                f"24.6× FY2027E EPS",     "peer P/E + premium",   475.0),
]
for lbl, met, mult, px in rel_methods:
    is_adopted = "Adopted" in lbl; is_blended = "Blended" in lbl
    bg_ = YEL if is_adopted else (LGR if is_blended else WHT)
    bold_ = is_adopted or is_blended
    label(ws, r, 2, lbl, bold=bold_, bg=bg_)
    cell(ws, r, 3, met,  halign="left", bg=bg_)
    cell(ws, r, 4, mult, halign="left", bg=bg_)
    c = money(ws, r, 5, px, bg=bg_, bold=bold_)
    c.number_format = '"$"#,##0.00'
    if is_adopted: c.font = Font(bold=True, color=GRN, name="Calibri")
    r += 1

r += 1
merge_hdr(ws, r, 2, 7, "SUM-OF-THE-PARTS (SOTP)", bg=BLU); r += 1
sotp_hdrs = ["Segment", "Revenue ($B)", "EBITDA ($B)", "Margin", "Peer Multiple", "Segment EV ($B)"]
for i, h in enumerate(sotp_hdrs):
    hdr(ws, r, 2+i, h)
r += 1

semi_rev = 33.9; semi_em = 0.55; semi_ebitda = round(semi_rev*semi_em, 1); semi_mult = 22.0; semi_ev = round(semi_ebitda*semi_mult, 1)
sw_rev   = 30.0; sw_em   = 0.72; sw_ebitda   = round(sw_rev  *sw_em,   1); sw_mult   = 25.0; sw_ev   = round(sw_ebitda  *sw_mult,   1)
total_seg = semi_ev + sw_ev
corp_cap  = -5.0
sotp_ev_  = total_seg + corp_cap
sotp_eq   = sotp_ev_ - net_debt_
sotp_px   = round(sotp_eq / shares, 2)

for row_ in [
    ("Semiconductor Solutions", semi_rev, semi_ebitda, semi_em, f"{semi_mult}× EV/EBITDA", semi_ev),
    ("Infrastructure Software", sw_rev,   sw_ebitda,   sw_em,   f"{sw_mult}× EV/EBITDA",  sw_ev),
]:
    lbl_, rv_, eb_, em_, ml_, ev_ = row_
    label(ws, r, 2, lbl_)
    money(ws, r, 3, rv_); money(ws, r, 4, eb_)
    pct(ws, r, 5, em_)
    cell(ws, r, 6, ml_, halign="left")
    money(ws, r, 7, ev_, bold=True)
    r += 1

sotp_bridge = [
    ("Total Segment EV ($B)",     total_seg),
    ("− Corporate costs ($B)",    corp_cap),
    ("− Net Debt ($B)",          -net_debt_),
    ("SOTP Equity Value ($B)",    sotp_eq),
    ("Diluted Shares (B)",        shares),
    ("→ SOTP Implied Price",      sotp_px),
    ("vs. Current $385.73",       f"{(sotp_px/385.73-1)*100:.1f}%"),
]
for lbl_, val_ in sotp_bridge:
    is_implied = "→" in lbl_
    label(ws, r, 2, lbl_, bold=is_implied)
    if isinstance(val_, str):
        cell(ws, r, 3, val_, halign="left", bold=True,
             fg=RED if "-" in val_ else GRN)
    else:
        c = money(ws, r, 3, val_, bold=is_implied)
        if is_implied:
            c.font = Font(bold=True, color=RED, name="Calibri")  # SOTP < current
            c.number_format = '"$"#,##0.00'
    r += 1

r += 1
cell(ws, r, 2, "Note: SOTP uses FY2025 segment EBITDAs — current value trap check, not forward growth valuation.",
     halign="left", italic=True, bg=LGR)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 7: Scenarios
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Scenarios")
set_col_widths(ws, [4, 35, 20, 20, 20, 20, 20, 4])
merge_hdr(ws, 1, 1, 8, "AVGO — Bull / Base / Bear Scenarios & Prob-Weighted Synthesis", bg=NAV, sz=12)

r = 3
merge_hdr(ws, r, 2, 8, "DCF SCENARIO ANALYSIS", bg=BLU); r += 1
scen_hdrs = ["Scenario", "WACC", "Terminal g", "TV Method", "Implied Price", "vs. Current"]
for i, h in enumerate(scen_hdrs):
    hdr(ws, r, 2+i, h)
r += 1

dcf_scens = [
    ("Bull",                 0.105, 0.030, "Gordon TV",                   px_gordon+63,   GRN),
    ("Base (Gordon TV)",     0.1218,0.025, "Gordon TV",                   px_gordon,      AMB),
    ("Base+ (blended TV)",   0.1218,0.025, "Gordon + 18× EBITDA exit",    px_blend,       AMB),
    ("Bear",                 0.140, 0.015, "Gordon TV; FCF 45% below base",94.0,          RED),
]
for name, w, g, tv, px, col in dcf_scens:
    label(ws, r, 2, name, bold=True)
    pct(ws, r, 3, w)
    pct(ws, r, 4, g)
    cell(ws, r, 5, tv, halign="left")
    c = money(ws, r, 6, px); c.number_format = '"$"#,##0.00'
    c.font = Font(bold=True, color=col, name="Calibri")
    diff = round((px / 385.73 - 1) * 100, 1)
    c2 = cell(ws, r, 7, f"{diff:+.1f}%", halign="center",
              bg=GRN if diff > 0 else RED)
    c2.font = Font(bold=True, color=WHT if diff < 0 else WHT, name="Calibri")
    c2.fill = fill(GRN if diff > 0 else RED)
    r += 1

r += 1
merge_hdr(ws, r, 2, 8, "PROBABILITY-WEIGHTED SYNTHESIS", bg=BLU); r += 1
pw_hdrs = ["Method", "Weight", "Implied Price", "Weighted Contribution", "Note"]
for i, h in enumerate(pw_hdrs):
    hdr(ws, r, 2+i, h)
r += 1

pw_rows = [
    ("DCF — Prob-Weighted",             0.35, 241,  84,  "30% bull + 45% base+ + 25% bear"),
    ("Relative — NTM P/E blend",        0.35, 389, 136,  "24.0× peer median × ~$16.2 NTM EPS"),
    ("Relative — FY2027E P/E",          0.20, 464,  93,  "24.0× × $19.321 FY27E EPS"),
    ("SOTP (conservative)",             0.10, 189,  19,  "FY25 segment EBITDAs; growth not captured"),
]
total_w = 0; total_px = 0
for name, w, px, wt_px, note in pw_rows:
    label(ws, r, 2, name)
    pct(ws, r, 3, w)
    c = money(ws, r, 4, px); c.number_format = '"$"#,##0'
    c2 = money(ws, r, 5, wt_px); c2.number_format = '"$"#,##0'
    cell(ws, r, 6, note, halign="left")
    total_w += w; total_px += wt_px
    r += 1

label(ws, r, 2, "Blended Intrinsic Value", bold=True, bg=YEL)
pct(ws, r, 3, total_w, bold=True)
c = money(ws, r, 4, "", bold=True, bg=YEL)
c = money(ws, r, 5, total_px, bold=True, bg=YEL)
c.number_format = '"$"#,##0'; c.font = Font(bold=True, color=AMB, name="Calibri", size=11)
r += 1
label(ws, r, 2, "Adopted Price Target", bold=True, bg=YEL)
c = money(ws, r, 4, 475, bold=True, bg=YEL)
c.number_format = '"$"#,##0'; c.font = Font(bold=True, color=GRN, name="Calibri", size=12)
c2 = cell(ws, r, 6, "24.6× FY2027E EPS $19.32; peer P/E + growth premium", halign="left", bg=YEL)
r += 2

merge_hdr(ws, r, 2, 8, "12-MONTH PRICE SCENARIO OUTCOMES", bg=BLU); r += 1
out_hdrs = ["Scenario", "Probability", "12-Month Target", "Weighted", "Key Assumption"]
for i, h in enumerate(out_hdrs):
    hdr(ws, r, 2+i, h)
r += 1

outcomes = [
    ("Bull", 0.35, 560, 196, "AI ramp intact, VMware retention >85%, FY27 $185–200B"),
    ("Base", 0.45, 460, 207, "Steady execution, FY27 tracks consensus ~$170B"),
    ("Bear", 0.20, 210,  42, "Hyperscaler XPU defection or VMware churn >15%"),
]
for name, p, tgt, wt, note in outcomes:
    col = GRN if tgt > 385 else RED
    label(ws, r, 2, name, bold=True)
    pct(ws, r, 3, p)
    c = money(ws, r, 4, tgt); c.number_format = '"$"#,##0'
    c.font = Font(bold=True, color=col, name="Calibri")
    money(ws, r, 5, wt); ws.cell(r,5).number_format = '"$"#,##0'
    cell(ws, r, 6, note, halign="left")
    r += 1

label(ws, r, 2, "Expected Value (PW)", bold=True, bg=YEL)
c = money(ws, r, 4, 445, bold=True, bg=YEL)
c.number_format = '"$"#,##0'; c.font = Font(bold=True, color=GRN, name="Calibri", size=11)
cell(ws, r, 6, "+15.4% vs current $385.73", halign="left", bg=YEL)

# ═════════════════════════════════════════════════════════════════════════════
# Sheet 8: Earnings & Estimates
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Earnings & Estimates")
set_col_widths(ws, [4, 35, 20, 20, 20, 20, 4])
merge_hdr(ws, 1, 1, 7, "AVGO — Earnings History, Estimates & EPS Trend (yfinance)", bg=NAV, sz=12)

r = 3
merge_hdr(ws, r, 2, 7, "EARNINGS BEAT HISTORY (yfinance earnings_history)", bg=BLU); r += 1
for i, h in enumerate(["Quarter", "EPS Estimate", "EPS Actual", "Surprise %", "Beat/Miss"]):
    hdr(ws, r, 2+i, h)
r += 1
beats = [
    ("Q3 FY2025 (Jul 31, 2025)", 1.663, 1.690, 0.016),
    ("Q4 FY2025 (Oct 31, 2025)", 1.868, 1.950, 0.0438),
    ("Q1 FY2026 (Jan 31, 2026)", 2.023, 2.050, 0.0132),
    ("Q2 FY2026 (Apr 30, 2026)", 2.397, 2.440, 0.0178),
]
for qtr, est, act, surp in beats:
    label(ws, r, 2, qtr)
    money(ws, r, 3, est); ws.cell(r,3).number_format = '"$"0.000'
    money(ws, r, 4, act); ws.cell(r,4).number_format = '"$"0.000'
    c = pct(ws, r, 5, surp, bold=True); c.font = Font(bold=True, color=GRN, name="Calibri")
    cell(ws, r, 6, "✓ Beat", bold=True, bg="D4EDDA")
    ws.cell(r,6).font = Font(bold=True, color=GRN, name="Calibri")
    r += 1

r += 1
merge_hdr(ws, r, 2, 7, "CONSENSUS ESTIMATES (yfinance revenue_estimate, earnings_estimate)", bg=BLU); r += 1
for i, h in enumerate(["Period", "Revenue Avg", "Revenue Low", "Revenue High", "EPS Avg", "EPS Low/High"]):
    hdr(ws, r, 2+i, h)
r += 1
ests = [
    ("Q3 FY2026 (0q)",      29.44, 29.21, 29.61, 3.237, "3.08–3.33"),
    ("Q4 FY2026 (+1q)",     34.86, 33.43, 36.16, 3.869, "3.60–4.12"),
    ("FY2026E (0y)",       106.01,104.33,109.13, 11.615,"11.23–12.14"),
    ("FY2027E (+1y)",      170.49,114.72,200.12, 19.321,"16.22–22.01"),
]
for period, ra, rl, rh, ea, eh in ests:
    label(ws, r, 2, period, bold=("2027" in period))
    for col, val in [(3,ra),(4,rl),(5,rh)]:
        c = money(ws, r, col, val); c.number_format = '"$"#,##0.0"B"'
    c = money(ws, r, 6, ea); c.number_format = '"$"0.000'
    cell(ws, r, 7, eh, halign="center")
    r += 1

r += 1
merge_hdr(ws, r, 2, 7, "EPS REVISION TREND — FY2027E (yfinance eps_trend +1y, eps_revisions)", bg=BLU); r += 1
for i, h in enumerate(["Period", "90d Ago", "60d Ago", "30d Ago", "7d Ago", "Current"]):
    hdr(ws, r, 2+i, h)
r += 1
label(ws, r, 2, "FY2027E EPS Consensus")
for col, val in [(3,17.31),(4,17.81),(5,18.08),(6,18.40),(7,19.321)]:
    c = money(ws, r, col, val); c.number_format = '"$"0.000'
    c.font = Font(bold=(col==7), color=GRN if col==7 else "000000", name="Calibri")
r += 1
label(ws, r, 2, "90-day EPS revision: +11.6% ($17.31 → $19.32)", bold=True, bg=LGR)

r += 2
merge_hdr(ws, r, 2, 7, "EPS REVISION BREADTH (yfinance eps_revisions, last 30 days)", bg=BLU); r += 1
for i, h in enumerate(["Period", "Up (30d)", "Down (30d)", "Ratio", "Signal"]):
    hdr(ws, r, 2+i, h)
r += 1
revisions = [
    ("Current Quarter (0q)",  4, 5, 0.44, "⚠ Bearish"),
    ("Next Quarter (+1q)",    3, 5, 0.38, "⚠ Bearish"),
    ("Current Year (0y)",     6, 5, 0.55, "→ Neutral"),
    ("Next Year (+1y)",      10, 3, 0.77, "✓ Bullish"),
]
for period, up, dn, ratio, sig in revisions:
    label(ws, r, 2, period)
    cell(ws, r, 3, up,    halign="center", fg=GRN)
    cell(ws, r, 4, dn,    halign="center", fg=RED)
    c = money(ws, r, 5, ratio); c.number_format = '0.00'
    col = GRN if ratio >= 0.60 else (RED if ratio < 0.50 else AMB)
    cell(ws, r, 6, sig, halign="center", bold=True, fg=col)
    r += 1

# ═════════════════════════════════════════════════════════════════════════════
# Finalise & save
# ═════════════════════════════════════════════════════════════════════════════
# Set active sheet to Scenarios
wb.active = wb["Scenarios"]

wb.save(OUT)
print(f"Excel written: {OUT}  ({os.path.getsize(OUT):,} bytes)")
